# collectors/excel_builder.py
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def build_ppc_excel_report(output_path, query_data=None):
    """
    Builds a professional Google Ads report with stylized Excel sheets.
    Sheet "Assets" is placed first after the "Alerts" ("Алерти") sheet.
    """
    if query_data is None:
        query_data = {}

    # Create excel writer
    writer = pd.ExcelWriter(output_path, engine='openpyxl')
    
    # 1. Alerts sheet data (Алерти)
    alerts_df = pd.DataFrame(query_data.get('alerts', [
        {"Акаунт": "Perfume Shop Kyiv", "Критичність": "🚨 Критична", "Опис": "Бюджет закінчиться через < 24 години", "Рекомендована дія": "Поповнити баланс на 10,000 UAH"},
        {"Акаунт": "B2B Tech Solutions", "Критичність": "⚠ Попередження", "Опис": "Різке падіння CTR (-45%) в кампанії B2B_Search", "Рекомендована дія": "Перевірити працездатність цільової сторінки"}
    ]))
    alerts_df.to_excel(writer, sheet_name="Алерти", index=False)
    
    # 2. Assets sheet (MUST be first after Alerts sheet)
    assets_df = pd.DataFrame(query_data.get('assets', [
        {
            "Кампанія": "Perfume Shop Kyiv - Пошук Бренду",
            "Група оголошень": "Брендові Ключі",
            "Тип активу": "HEADLINE (Заголовок)",
            "Текст": "Купити Духи Оригінал",
            "Performance Label": "BEST (Найкращий)",
            "Impressions": 4820,
            "Clicks": 591,
            "Pinned": "UNPINNED"
        },
        {
            "Кампанія": "Perfume Shop Kyiv - Пошук Бренду",
            "Група оголошень": "Брендові Ключі",
            "Тип активу": "DESCRIPTION (Опис)",
            "Текст": "Офіційний дистриб'ютор парфумерії в Україні. Безкоштовна доставка.",
            "Performance Label": "GOOD (Хороший)",
            "Impressions": 3120,
            "Clicks": 341,
            "Pinned": "PINNED_TO_DESCRIPTION_1"
        },
        {
            "Кампанія": "B2B Tech Solutions - Search Main",
            "Група оголошень": "Послуги Інтеграції",
            "Тип активу": "SITELINK (Посилання)",
            "Текст": "Акційні пропозиції",
            "Performance Label": "LEARNING (Навчання)",
            "Impressions": 850,
            "Clicks": 65,
            "Pinned": "UNPINNED"
        },
        {
            "Кампанія": "B2B Tech Solutions - Search Main",
            "Група оголошень": "Послуги Інтеграції",
            "Тип активу": "CALLOUT (Уточнення)",
            "Текст": "Гарантія якості 3 роки",
            "Performance Label": "LOW (Слабкий)",
            "Impressions": 1240,
            "Clicks": 89,
            "Pinned": "UNPINNED"
        }
    ]))
    # Reorder columns explicitly to match user instructions:
    # Campaign | Ad Group | Asset Type | Text | Performance Label | Impressions | Clicks | Pinned
    assets_columns = [
        "Кампанія", "Група оголошень", "Тип активу", "Текст", "Performance Label", "Impressions", "Clicks", "Pinned"
    ]
    assets_df = assets_df.reindex(columns=assets_columns)
    assets_df.to_excel(writer, sheet_name="Assets", index=False)
    
    # 3. Campaigns sheet data
    campaigns_df = pd.DataFrame(query_data.get('campaigns', [
        {"ID": "123456789", "Кампанія": "Perfume Shop Kyiv - Пошук Бренду", "Статус": "ENABLED", "Тип": "SEARCH", "Impressions": 12450, "Clicks": 1420, "Cost": "15,400.50 UAH"},
        {"ID": "987654321", "Кампанія": "B2B Tech Solutions - Search Main", "Статус": "ENABLED", "Тип": "PERFORMANCE_MAX", "Impressions": 45100, "Clicks": 3490, "Cost": "34,120.20 UAH"}
    ]))
    campaigns_df.to_excel(writer, sheet_name="Кампанії", index=False)
    
    # Close first to compile files
    writer.close()
    
    # Load workbook for detailed styling
    wb = openpyxl.load_workbook(output_path)
    
    # Styles configuration
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")  # Dark Blue
    zebra_fill = PatternFill(start_color="F2F5F8", end_color="F2F5F8", fill_type="solid")   # Very light Blue/Gray
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Arial", size=10)
    border_side = Side(border_style="thin", color="D9D9D9")
    data_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    
    # Reorder sheets to put "Assets" first after "Алерти"
    sheet_order = ["Алерти", "Assets", "Кампанії"]
    wb._sheets = [wb[s] for s in sheet_order if s in wb.sheetnames]
    
    for name in wb.sheetnames:
        ws = wb[name]
        ws.views.sheetView[0].showGridLines = True
        
        # Style Header Row
        for col_num in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = data_border
        
        ws.row_dimensions[1].height = 28
        
        # Style Data Rows
        for row_num in range(2, ws.max_row + 1):
            is_zebra = (row_num % 2 == 0)
            ws.row_dimensions[row_num].height = 20
            for col_num in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.font = data_font
                cell.border = data_border
                
                # Zebra striping
                if is_zebra:
                    cell.fill = zebra_fill
                
                # Format alignment and formats based on type
                val = cell.value
                if isinstance(val, (int, float)):
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    # Format as currency or integers
                    header_name = str(ws.cell(row=1, column=col_num).value)
                    if "Spend" in header_name or "Cost" in header_name:
                        cell.number_format = '#,##0.00'
                    else:
                        cell.number_format = '#,##0'
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                    
        # Autofit columns
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or '')
                if len(val_str) > max_len:
                    max_len = len(val_str)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 14)
            
    wb.save(output_path)
    wb.close()
    print(f"Professional Excel report created at: {output_path}")

if __name__ == "__main__":
    build_ppc_excel_report("ppc_analytics_report.xlsx")
